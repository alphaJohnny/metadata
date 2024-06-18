import os
import subprocess
from openpyxl import load_workbook
from datetime import datetime

def get_excel_metadata(file_path):
    """Retrieve metadata from an Excel file."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    
    wb = load_workbook(file_path)
    metadata = wb.properties
    
    metadata_dict = {
        'Title': metadata.title,
        'Subject': metadata.subject,
        'Creator': metadata.creator,
        'Keywords': metadata.keywords,
        'Description': metadata.description,
        'Last Modified By': metadata.lastModifiedBy,
        'Revision': metadata.revision,
        'Created': metadata.created,
        'Modified': metadata.modified,
        'Category': metadata.category,
        'Content Status': metadata.contentStatus,
        'Language': metadata.language,
        'Identifier': metadata.identifier
    }
    
    return metadata_dict

def set_excel_metadata(file_path, new_metadata):
    """Set new metadata to an Excel file."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    
    wb = load_workbook(file_path)
    
    # Set new properties
    wb.properties.title = new_metadata.get('Title', wb.properties.title)
    wb.properties.subject = new_metadata.get('Subject', wb.properties.subject)
    wb.properties.creator = new_metadata.get('Creator', wb.properties.creator)
    wb.properties.keywords = new_metadata.get('Keywords', wb.properties.keywords)
    wb.properties.description = new_metadata.get('Description', wb.properties.description)
    wb.properties.lastModifiedBy = new_metadata.get('Last Modified By', wb.properties.lastModifiedBy)
    
    # Increment revision number automatically
    wb.properties.revision = str(int(wb.properties.revision) + 1) if wb.properties.revision else '1'
    
    wb.properties.created = new_metadata.get('Created', wb.properties.created)
    wb.properties.modified = new_metadata.get('Modified', wb.properties.modified)
    wb.properties.category = new_metadata.get('Category', wb.properties.category)
    wb.properties.contentStatus = new_metadata.get('Content Status', wb.properties.contentStatus)
    wb.properties.language = new_metadata.get('Language', wb.properties.language)
    wb.properties.identifier = new_metadata.get('Identifier', wb.properties.identifier)
    
    wb.save(file_path)
    return get_excel_metadata(file_path)

def prompt_for_metadata(existing_metadata):
    """Prompt user for new metadata."""
    new_metadata = {}
    
    for key, value in existing_metadata.items():
        print(f"{key}: {value}")
        if key in ['Created', 'Modified']:
            response = input(f"Do you want to change the {key}? (current: {value}) (yes/no): ").strip().lower()
            if response == 'yes':
                date_str = input(f"Enter the new {key} date (YYYY-MM-DD HH:MM:SS): ").strip()
                try:
                    new_metadata[key] = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    print(f"Invalid date format for {key}. Keeping the existing value.")
                    new_metadata[key] = value
            else:
                new_metadata[key] = value
        else:
            response = input(f"Enter new value for {key} (current: {value}): ").strip()
            new_metadata[key] = response if response else value
    
    # Automatically handle the revision increment
    if 'Revision' in existing_metadata:
        print(f"Revision: {existing_metadata['Revision']}")
        new_metadata['Revision'] = str(int(existing_metadata['Revision']) + 1) if existing_metadata['Revision'] else '1'
        print(f"New Revision: {new_metadata['Revision']}")

    return new_metadata

def update_file_system_dates(file_path, created, modified):
    """Update file system creation and modification dates to match Excel metadata."""
    if created:
        created_timestamp = datetime.timestamp(created)
        subprocess.run(['touch', '-t', datetime.fromtimestamp(created_timestamp).strftime('%Y%m%d%H%M.%S'), file_path])

    if modified:
        modified_timestamp = datetime.timestamp(modified)
        subprocess.run(['touch', '-mt', datetime.fromtimestamp(modified_timestamp).strftime('%Y%m%d%H%M.%S'), file_path])

# Example usage:
file_path = 'Timesheet record CSCIL.xlsx'

# Get current metadata
current_metadata = get_excel_metadata(file_path)
print("Current Metadata:")
for key, value in current_metadata.items():
    print(f"{key}: {value}")

# Ask if user wants to change metadata
change_metadata = input("Do you want to change the metadata? (yes/no): ").strip().lower()

if change_metadata == 'yes':
    new_metadata = prompt_for_metadata(current_metadata)
    updated_metadata = set_excel_metadata(file_path, new_metadata)
    print("\nUpdated Metadata:")
    for key, value in updated_metadata.items():
        print(f"{key}: {value}")
    
    # Update file system dates to match Excel metadata
    update_file_system_dates(file_path, updated_metadata['Created'], updated_metadata['Modified'])
else:
    print("Metadata not changed.")
