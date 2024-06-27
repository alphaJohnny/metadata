import os
import hashlib
import platform
import tkinter
from tkinter import Tk, filedialog, Label, Button, ttk, messagebox, font, simpledialog
from openpyxl import load_workbook
from datetime import datetime
from PyPDF2 import PdfReader
from PIL import Image
from PIL.ExifTags import TAGS
from tkcalendar import DateEntry
from docx import Document
from docx.opc.coreprops import CoreProperties
import tkinter as tk
from tkinter import ttk
from Foundation import NSDate, NSFileManager, NSURL
from win32_setctime import setctime
from openpyxl.packaging.core import DocumentProperties
from dateutil import parser as date_parser
import pytz


class ProgressDialog:
    def __init__(self, parent, title="Processing"):
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.geometry("300x100")
        self.top.transient(parent)
        self.top.grab_set()

        self.progress = ttk.Progressbar(self.top, mode="indeterminate", length=280)
        self.progress.pack(pady=20)

        self.label = ttk.Label(self.top, text="Please wait...")
        self.label.pack()

    def start(self):
        self.progress.start()

    def stop(self):
        self.progress.stop()
        self.top.destroy()

class DateTimeDialog(simpledialog.Dialog):
    def __init__(self, parent, title, initial_date=None):
        self.initial_date = initial_date or datetime.now()
        super().__init__(parent, title)

    def body(self, master):
        self.calendar = DateEntry(master, width=12, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='y-mm-dd',
                                  year=self.initial_date.year, month=self.initial_date.month, day=self.initial_date.day)
        self.calendar.grid(row=0, column=0, padx=5, pady=5, sticky="we")

        self.time_frame = ttk.Frame(master)
        self.time_frame.grid(row=1, column=0, padx=5, pady=5, sticky="we")

        self.hour = ttk.Spinbox(self.time_frame, from_=0, to=23, width=3, format="%02.0f")
        self.hour.grid(row=0, column=0)
        ttk.Label(self.time_frame, text=":").grid(row=0, column=1)
        self.minute = ttk.Spinbox(self.time_frame, from_=0, to=59, width=3, format="%02.0f")
        self.minute.grid(row=0, column=2)

        self.hour.set(f"{self.initial_date.hour:02d}")
        self.minute.set(f"{self.initial_date.minute:02d}")

        return self.calendar

    def apply(self):
        date = self.calendar.get_date()
        hour = int(self.hour.get())
        minute = int(self.minute.get())
        self.result = datetime(date.year, date.month, date.day, hour, minute)

def parse_date(date_string):
    try:
        # Parse the date string, forcing it to UTC
        parsed_date = date_parser.parse(date_string)
        if parsed_date.tzinfo is not None:
            parsed_date = parsed_date.astimezone(pytz.UTC)
        else:
            parsed_date = pytz.UTC.localize(parsed_date)
        # Return the date without timezone info
        return parsed_date.replace(tzinfo=None)
    except ValueError:
        # If parsing fails, return None
        return None

def get_excel_metadata(file_path):
    """Retrieve metadata from an Excel file."""
    wb = load_workbook(file_path)
    metadata = wb.properties
    
    metadata_dict = {
        'Title': metadata.title,
        'Subject': metadata.subject,
        'Creator': metadata.creator,
        'Keywords': metadata.keywords,
        'Description': metadata.description,
        'Last Modified By': metadata.lastModifiedBy,
        'Revision': metadata.revision,
        'Created': metadata.created,
        'Modified': metadata.modified,
        'Category': metadata.category,
        'Content Status': metadata.contentStatus,
        'Language': metadata.language,
        'Identifier': metadata.identifier
    }
    
    return metadata_dict

def get_pdf_metadata(file_path):
    """Retrieve metadata from a PDF file."""
    with open(file_path, 'rb') as f:
        reader = PdfReader(f)
        metadata = reader.metadata
    
    metadata_dict = {key[1:]: value for key, value in metadata.items()}
    return metadata_dict

def get_image_metadata(file_path):
    """Retrieve metadata from an image file."""
    image = Image.open(file_path)
    info = image._getexif()
    
    metadata_dict = {TAGS.get(tag): value for tag, value in info.items()} if info else {}
    return metadata_dict

def get_docx_metadata(file_path):
    """Retrieve metadata from a DOCX file."""
    doc = Document(file_path)
    core_props = doc.core_properties
    
    metadata_dict = {
        'Title': core_props.title,
        'Subject': core_props.subject,
        'Creator': core_props.author,
        'Keywords': core_props.keywords,
        'Description': core_props.comments,
        'Last Modified By': core_props.last_modified_by,
        'Revision': core_props.revision,
        'Created': core_props.created,
        'Modified': core_props.modified,
        'Category': core_props.category,
        'Content Status': core_props.content_status,
        'Language': core_props.language,
        'Identifier': core_props.identifier
    }
    
    return metadata_dict

def update_file_system_dates(file_path, created, modified):
    """Update file system creation and modification dates to match file metadata."""
    if modified:
        os.utime(file_path, (modified.timestamp(), modified.timestamp()))
    
    if created:
        if platform.system() == 'Darwin':  # macOS
            try:
                from Foundation import NSDate, NSFileManager, NSURL
            except ImportError:
                pass  # Handle the case when pyobjc is not installed
        elif platform.system() == 'Windows':
            try:
                from win32_setctime import setctime
            except ImportError:
                pass  # Handle the case when win32-setctime is not installed

def calculate_hash(file_path, algorithm='sha256'):
    """Calculate and return the hash of the file."""
    hash_func = hashlib.new(algorithm)
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b''):
            hash_func.update(chunk)
    return hash_func.hexdigest()

def compare_to_industry_standards(metadata):
    """Compare metadata to industry standards and identify potential issues."""
    issues = {}
    if metadata.get('Creator') is None:
        issues['Creator'] = 'Missing creator'
    if metadata.get('Revision') is None or not str(metadata['Revision']).isdigit():
        issues['Revision'] = 'Invalid revision number'
    
    return issues

def display_metadata(metadata, issues):
    """Display metadata in a table with potential issues."""
    for item in tree.get_children():
        tree.delete(item)
    for key, value in metadata.items():
        print(f"Displaying metadata: {key} = {value}")  # Debug print
        tree.insert("", "end", values=(key, str(value), issues.get(key, '')))
    
    # Debug print of all treeview items
    print("All treeview items:")
    for item in tree.get_children():
        print(tree.item(item, 'values'))

def on_select_file():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("PDF files", "*.pdf"), ("Image files", "*.jpg *.jpeg *.png *.gif"), ("Word files", "*.docx")])
        if not file_path:
            messagebox.showwarning("No Selection", "No file was selected.")
            return

        if file_path.endswith('.xlsx'):
            metadata = get_excel_metadata(file_path)
        elif file_path.endswith('.pdf'):
            metadata = get_pdf_metadata(file_path)
        elif file_path.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
            metadata = get_image_metadata(file_path)
        elif file_path.endswith('.docx'):
            metadata = get_docx_metadata(file_path)
        else:
            messagebox.showerror("Unsupported File", "The selected file type is not supported.")
            return

        issues = compare_to_industry_standards(metadata)
        display_metadata(metadata, issues)
        file_label.config(text=f"Selected file: {file_path}")
        global selected_file
        selected_file = file_path

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while selecting the file: {str(e)}")


def on_double_click(event):
    """Handle double click event to edit metadata."""
    item = tree.identify('item', event.x, event.y)
    column = tree.identify_column(event.x)

    if column == '#2':
        key = tree.item(item, 'values')[0]
        value = tree.item(item, 'values')[1]

        if key in ['Created', 'Modified']:
            try:
                initial_date = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                initial_date = datetime.now()
            
            root.attributes('-topmost', False)
            dialog = DateTimeDialog(root, f"Select {key} Date and Time", initial_date)
            root.attributes('-topmost', True)
            if dialog.result:
                new_value = dialog.result.strftime("%Y-%m-%d %H:%M:%S")
                tree.set(item, column=1, value=new_value)
        else:
            x, y, width, height = tree.bbox(item, column)
            
            entry_edit = ttk.Entry(tree)
            entry_edit.insert(0, value)
            entry_edit.select_range(0, 'end')

            entry_edit.place(x=x, y=y, width=width, height=height)

            def save_edit(event=None):
                """Save the edited value to the treeview."""
                new_value = entry_edit.get()
                tree.set(item, column=1, value=new_value)
                entry_edit.destroy()

            def cancel_edit(event=None):
                """Cancel the edit without saving."""
                entry_edit.destroy()

            entry_edit.bind('<Return>', save_edit)
            entry_edit.bind('<Escape>', cancel_edit)
            entry_edit.bind('<FocusOut>', cancel_edit)
            
            # Use after() to schedule the focus_force() call
            root.after(100, entry_edit.focus_force)


def on_save():
    progress = ProgressDialog(root, "Saving Metadata")
    progress.start()
    root.update()
    
    if not selected_file:
        messagebox.showerror("Error", "No file selected")
        progress.stop()
        return

    new_metadata = {}
    for child in tree.get_children():
        key = tree.item(child, 'values')[0]
        value = tree.item(child, 'values')[1]
        new_metadata[key] = value
    
    try:
        if selected_file.endswith('.xlsx'):
            wb = load_workbook(selected_file)
            
            # Create a new DocumentProperties object
            new_props = DocumentProperties()
            
            # Update all properties
            new_props.title = new_metadata.get('Title', wb.properties.title)
            new_props.subject = new_metadata.get('Subject', wb.properties.subject)
            new_props.creator = new_metadata.get('Creator', wb.properties.creator)
            new_props.keywords = new_metadata.get('Keywords', wb.properties.keywords)
            new_props.description = new_metadata.get('Description', wb.properties.description)
            new_props.lastModifiedBy = new_metadata.get('Last Modified By', wb.properties.lastModifiedBy)
            
            # Update revision number
            revision = new_metadata.get('Revision', '0')
            new_props.revision = str(int(revision)) if revision.isdigit() else '0'
            
            new_props.category = new_metadata.get('Category', wb.properties.category)
            new_props.contentStatus = new_metadata.get('Content Status', wb.properties.contentStatus)
            new_props.language = new_metadata.get('Language', wb.properties.language)
            new_props.identifier = new_metadata.get('Identifier', wb.properties.identifier)
            
            created = new_metadata.get('Created')
            modified = new_metadata.get('Modified')

            if created:
                try:
                    created = datetime.strptime(created, "%Y-%m-%d %H:%M:%S")
                    new_props.created = created
                except ValueError as e:
                    messagebox.showerror("Error", f"Incorrect date format for 'Created': {e}")
                    progress.stop()
                    return

            if modified:
                try:
                    modified = datetime.strptime(modified, "%Y-%m-%d %H:%M:%S")
                    new_props.modified = modified
                except ValueError as e:
                    messagebox.showerror("Error", f"Incorrect date format for 'Modified': {e}")
                    progress.stop()
                    return
            
            # Assign the new properties to the workbook
            wb.properties = new_props
            
            wb.save(selected_file)
            
            update_file_system_dates(selected_file, created, modified)

        elif selected_file.endswith('.docx'):
            doc = Document(selected_file)
            core_props = doc.core_properties
            
            core_props.title = new_metadata.get('Title', core_props.title)
            core_props.subject = new_metadata.get('Subject', core_props.subject)
            core_props.author = new_metadata.get('Creator', core_props.author)
            core_props.keywords = new_metadata.get('Keywords', core_props.keywords)
            core_props.comments = new_metadata.get('Description', core_props.comments)
            core_props.last_modified_by = new_metadata.get('Last Modified By', core_props.last_modified_by)
            
            # Handle revision number
            revision = new_metadata.get('Revision', '0')
            core_props.revision = int(revision) if revision.isdigit() else 0
            
            core_props.category = new_metadata.get('Category', core_props.category)
            core_props.content_status = new_metadata.get('Content Status', core_props.content_status)
            core_props.language = new_metadata.get('Language', core_props.language)
            core_props.identifier = new_metadata.get('Identifier', core_props.identifier)
            
            # Handle dates
            created = parse_date(new_metadata.get('Created'))
            modified = parse_date(new_metadata.get('Modified'))

            if created:
                core_props.created = created
            if modified:
                core_props.modified = modified
            
            doc.save(selected_file)
            
            update_file_system_dates(selected_file, created, modified)
            
        else:
            messagebox.showerror("Unsupported File", "Saving metadata is only supported for Excel and Word files in this version.")
            progress.stop()
            return

        file_hash = calculate_hash(selected_file)
        hash_label.config(text=f"File Hash (SHA-256): {file_hash}")

        progress.stop()
        messagebox.showinfo("Success", "File metadata updated successfully")
    except Exception as e:
        progress.stop()
        messagebox.showerror("Error", f"An error occurred while saving the metadata: {str(e)}")
        print(f"Error details: {e}")

# GUI setup
root = Tk()
root.title("Forensic Audit Tool")
root.geometry("800x600")

# Mac-specific adjustments
if platform.system() == 'Darwin':  # Darwin is the system name for macOS
    try:
        from tkmacosx import Button as MacButton
        Button = MacButton  # Use tkmacosx Button if available
    except ImportError:
        pass  # Fall back to standard tkinter Button if tkmacosx is not installed
    
    # Adjust font size for better readability on Mac
    default_font = font.nametofont("TkDefaultFont")
    default_font.configure(size=12)
    root.option_add("*Font", default_font)

    # Try to set native macOS appearance
    try:
        root.tk.call('::tk::unsupported::MacWindowStyle', 'useTheme', 'true')
    except tkinter.TclError:
        pass  # Ignore if the command is not available

selected_file = None

# File selection
file_label = Label(root, text="No file selected")
file_label.pack(pady=10)
select_button = Button(root, text="Select File", command=on_select_file)
select_button.pack(pady=10)

# Metadata table
style = ttk.Style()
style.configure("Treeview", rowheight=25)  # Adjust row height for better touch/click targets

columns = ('Property', 'Value', 'Issues')
tree = ttk.Treeview(root, columns=columns, show='headings', selectmode='browse')
tree.heading('Property', text='Property')
tree.heading('Value', text='Value')
tree.heading('Issues', text='Issues')
tree.column('Property', width=150)
tree.column('Value', width=300)
tree.column('Issues', width=200)
tree.pack(pady=20, fill='both', expand=True)

# Bind the double click event to the handler
tree.bind('<Double-1>', on_double_click)

# Save button
save_button = Button(root, text="Save Changes", command=on_save)
save_button.pack(pady=10)

# Hash label
hash_label = Label(root, text="File Hash (SHA-256): N/A")
hash_label.pack(pady=10)

root.mainloop()